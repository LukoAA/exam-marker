"""Excel export for a single validated MarkingReport (Phase 1 CLI use):
`python -m app.cli mark --pdf x.pdf --scheme scheme.txt --course "..." --out report.xlsx`.

The batch-level workbook described in PROJECT_SPEC.md's "EXCEL EXPORT"
section (Summary / Flags / Item Analysis sheets across many students) is a
Phase 2+ concern once there's a batch of scripts to summarize; this module
covers the single-script case only.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas import MarkingReport

_BOLD = Font(bold=True)


def _bold_last_row(ws: Worksheet) -> None:
    for cell in ws[ws.max_row]:
        cell.font = _BOLD


def _write_report_sheet(ws: Worksheet, report: MarkingReport) -> None:
    ws.title = "Report"

    ws.append(["Student Name", report.student_name or ""])
    ws.append(["Matric Number", report.matric_number or ""])
    ws.append(["Course Code", report.course_code])
    ws.append(["Needs Human Review", "YES" if report.needs_human_review else "NO"])
    if report.review_reasons:
        ws.append(["Review Reasons", "; ".join(report.review_reasons)])
    ws.append([])

    ws.append(["Question", "Max Marks", "Awarded", "Provisional"])
    _bold_last_row(ws)

    for question in report.questions:
        ws.append(
            [
                f"Q{question.question}",
                question.max_marks,
                question.awarded,
                "Yes" if question.provisional else "No",
            ]
        )
        ws.append(["", "Mark Point", "Decision", "Marks", "Evidence"])
        _bold_last_row(ws)
        for mark_point in question.mark_points:
            ws.append(
                ["", mark_point.scheme_point, mark_point.decision, mark_point.marks, mark_point.evidence]
            )
        ws.append(["", "Strengths", question.strengths])
        ws.append(["", "Missing points", question.missing_points])
        ws.append(["", "Errors", question.errors])
        ws.append([])

    ws.append(["Total Awarded", report.total_awarded])
    ws.append(["Total Possible", report.total_possible])
    ws.append(["Percentage", report.percentage])
    ws.append(["Grade", report.grade or ""])


def _write_flags_sheet(ws: Worksheet, report: MarkingReport) -> None:
    ws.title = "Flags"

    ws.append(["Type", "Page", "Question", "Detail", "Impact"])
    _bold_last_row(ws)

    for item in report.low_confidence_sections:
        ws.append(["Low Confidence", item.page, item.question, item.text, item.impact])

    for item in report.page_anomalies:
        ws.append(
            [
                f"Page Anomaly ({item.type})",
                item.page,
                ", ".join(item.affected_questions),
                item.detail,
                "",
            ]
        )

    for detail in report.identity_anomalies:
        ws.append(["Identity Anomaly", "", "", detail, ""])


def export_report_to_excel(report: MarkingReport, output_path: str | Path) -> Path:
    """Write one script's marking report to an .xlsx workbook: a "Report"
    sheet (per-question breakdown, mark points, totals, grade, review flag)
    and a "Flags" sheet (low-confidence and anomaly items)."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    _write_report_sheet(workbook.active, report)
    _write_flags_sheet(workbook.create_sheet(), report)
    workbook.save(output_path)

    return output_path

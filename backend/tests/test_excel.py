from openpyxl import load_workbook

from app.export.excel import export_report_to_excel
from app.schemas import MarkingReport

SAMPLE_REPORT = {
    "student_name": "Jane Doe",
    "matric_number": "CSC/2021/001",
    "course_code": "CSC301",
    "needs_human_review": True,
    "review_reasons": ["Question 3 answer not anticipated by the scheme"],
    "questions": [
        {
            "question": "1",
            "attempted": True,
            "max_marks": 10,
            "awarded": 7.5,
            "provisional": False,
            "mark_points": [
                {
                    "scheme_point": "1a — definition of osmosis (2 marks)",
                    "decision": "AWARDED",
                    "marks": 2,
                    "evidence": "student wrote: 'movement of water molecules...'",
                    "note": "",
                },
                {
                    "scheme_point": "1b — three characteristics @ 1 (3 marks)",
                    "decision": "PARTIAL",
                    "marks": 2,
                    "evidence": "student gave two of three characteristics",
                    "note": "missing the third characteristic",
                },
            ],
            "strengths": "Clear definition and good terminology.",
            "missing_points": "Third characteristic omitted.",
            "errors": "None.",
            "legibility_flags": [],
        },
        {
            "question": "2",
            "attempted": True,
            "max_marks": 5,
            "awarded": 5,
            "provisional": False,
            "mark_points": [
                {
                    "scheme_point": "2a — correct formula (5 marks)",
                    "decision": "AWARDED",
                    "marks": 5,
                    "evidence": "student wrote the correct formula",
                    "note": "",
                }
            ],
            "strengths": "Fully correct.",
            "missing_points": "None.",
            "errors": "None.",
            "legibility_flags": [],
        },
    ],
    "mcq_section": {
        "present": False,
        "answer_string": "",
        "correct": 0,
        "wrong": 0,
        "blank": 0,
        "ambiguous": 0,
        "score": 0,
    },
    "page_anomalies": [
        {
            "page": 4,
            "type": "rotated",
            "detail": "Page scanned upside down; content read correctly regardless.",
            "affected_questions": ["3"],
        }
    ],
    "identity_anomalies": ["Handwriting on page 3 differs noticeably from pages 1-2."],
    "total_awarded": 12.5,
    "total_possible": 15,
    "percentage": 83.33,
    "grade": "A",
    "low_confidence_sections": [
        {
            "page": 2,
            "question": "3",
            "text": "[ILLEGIBLE: 4 words]",
            "impact": "up to 2 marks undetermined",
        }
    ],
    "overall_feedback": {
        "concepts_understood": "Osmosis fundamentals.",
        "weak_areas": "Detailed characteristics.",
        "topics_to_revise": "Cell membrane transport.",
        "summary": "Solid grasp of core concept, needs more depth.",
    },
}


def _all_rows(ws):
    """Row values with trailing Nones stripped — openpyxl pads every row out
    to the sheet's widest row when iterating, which would otherwise make
    exact-length comparisons against shorter appended rows fail."""
    rows = []
    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        while values and values[-1] is None:
            values.pop()
        rows.append(values)
    return rows


def test_export_report_to_excel_creates_readable_workbook(tmp_path):
    report = MarkingReport.model_validate(SAMPLE_REPORT)
    output_path = tmp_path / "report.xlsx"

    result_path = export_report_to_excel(report, output_path)

    assert result_path == output_path
    assert output_path.exists()

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Report", "Flags"]


def test_report_sheet_contains_header_and_totals(tmp_path):
    report = MarkingReport.model_validate(SAMPLE_REPORT)
    output_path = tmp_path / "report.xlsx"
    export_report_to_excel(report, output_path)

    ws = load_workbook(output_path)["Report"]
    rows = _all_rows(ws)

    assert ["Student Name", "Jane Doe"] in rows
    assert ["Matric Number", "CSC/2021/001"] in rows
    assert ["Course Code", "CSC301"] in rows
    assert ["Needs Human Review", "YES"] in rows
    assert ["Total Awarded", 12.5] in rows
    assert ["Total Possible", 15] in rows
    assert ["Percentage", 83.33] in rows
    assert ["Grade", "A"] in rows


def test_report_sheet_contains_question_and_mark_point_rows(tmp_path):
    report = MarkingReport.model_validate(SAMPLE_REPORT)
    output_path = tmp_path / "report.xlsx"
    export_report_to_excel(report, output_path)

    ws = load_workbook(output_path)["Report"]
    rows = _all_rows(ws)

    assert ["Q1", 10, 7.5, "No"] in rows
    assert ["Q2", 5, 5, "No"] in rows
    assert [
        None,
        "1a — definition of osmosis (2 marks)",
        "AWARDED",
        2,
        "student wrote: 'movement of water molecules...'",
    ] in rows
    assert [
        None,
        "1b — three characteristics @ 1 (3 marks)",
        "PARTIAL",
        2,
        "student gave two of three characteristics",
    ] in rows


def test_flags_sheet_contains_anomalies_and_low_confidence_items(tmp_path):
    report = MarkingReport.model_validate(SAMPLE_REPORT)
    output_path = tmp_path / "report.xlsx"
    export_report_to_excel(report, output_path)

    ws = load_workbook(output_path)["Flags"]
    rows = _all_rows(ws)

    assert rows[0] == ["Type", "Page", "Question", "Detail", "Impact"]
    assert [
        "Low Confidence",
        2,
        "3",
        "[ILLEGIBLE: 4 words]",
        "up to 2 marks undetermined",
    ] in rows
    assert [
        "Page Anomaly (rotated)",
        4,
        "3",
        "Page scanned upside down; content read correctly regardless.",
    ] in rows
    assert [
        "Identity Anomaly",
        None,
        None,
        "Handwriting on page 3 differs noticeably from pages 1-2.",
    ] in rows

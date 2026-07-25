import json
from pathlib import Path
from types import SimpleNamespace

import cv2
import numpy as np
import pytest
from openpyxl import load_workbook
from PIL import Image

from app import cli


def _make_lined_image(width: int = 300, height: int = 400) -> np.ndarray:
    image = np.full((height, width, 3), 255, dtype=np.uint8)
    y = 20
    while y + 12 < height - 20:
        cv2.rectangle(image, (20, y), (width - 20, y + 12), (0, 0, 0), thickness=-1)
        y += 22
    return image


@pytest.fixture
def sample_pdf(tmp_path):
    image = Image.fromarray(cv2.cvtColor(_make_lined_image(), cv2.COLOR_BGR2RGB))
    pdf_path = tmp_path / "script.pdf"
    image.save(pdf_path, "PDF")
    return pdf_path


@pytest.fixture
def scheme_file(tmp_path):
    path = tmp_path / "scheme.txt"
    path.write_text("Q1 (10 marks):\n  1a. Definition ..... 2 marks\n", encoding="utf-8")
    return path


def _valid_report_dict():
    return {
        "student_name": "Jane Doe",
        "matric_number": "CSC/2021/001",
        "course_code": "CSC301",
        "needs_human_review": False,
        "review_reasons": [],
        "questions": [
            {
                "question": "1",
                "attempted": True,
                "max_marks": 10,
                "awarded": 7,
                "provisional": False,
                "mark_points": [
                    {
                        "scheme_point": "1a — definition (2 marks)",
                        "decision": "AWARDED",
                        "marks": 2,
                        "evidence": "student wrote: '...'",
                        "note": "",
                    }
                ],
                "strengths": "Good.",
                "missing_points": "None.",
                "errors": "None.",
                "legibility_flags": [],
            }
        ],
        "mcq_section": {
            "present": False,
            "answer_string": "",
            "correct": 0,
            "wrong": 0,
            "blank": 0,
            "ambiguous": 0,
            "score": 0,
        },
        "page_anomalies": [],
        "identity_anomalies": [],
        "total_awarded": 7,
        "total_possible": 10,
        "percentage": 70,
        "grade": "B",
        "low_confidence_sections": [],
        "overall_feedback": {
            "concepts_understood": "…",
            "weak_areas": "…",
            "topics_to_revise": "…",
            "summary": "…",
        },
    }


def _full_response_text(report_dict: dict) -> str:
    json_block = f"```json\n{json.dumps(report_dict, indent=2)}\n```"
    return (
        "**Part A — JSON (for the software):**\n\n"
        f"{json_block}\n\n"
        "**Part B — Human-readable report:**\n\n"
        "=================================================\n"
        "STUDENT EXAMINATION REPORT\n"
        "=================================================\n"
        "TOTAL: 7/10\n"
    )


def _text_response(text: str) -> SimpleNamespace:
    return SimpleNamespace(content=[SimpleNamespace(type="text", text=text)])


class FakeMessages:
    def __init__(self, responses):
        self._responses = iter(responses)
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        return next(self._responses)


class FakeAnthropic:
    def __init__(self, responses):
        self.messages = FakeMessages(responses)


# ---------------------------------------------------------------------------
# argument parsing
# ---------------------------------------------------------------------------


def test_parser_accepts_required_arguments_and_applies_defaults():
    parser = cli.build_parser()

    args = parser.parse_args(
        [
            "mark",
            "--pdf",
            "script.pdf",
            "--scheme",
            "scheme.txt",
            "--course",
            "CSC301",
            "--total",
            "100",
            "--out",
            "report.json",
        ]
    )

    assert args.pdf == Path("script.pdf")
    assert args.scheme == Path("scheme.txt")
    assert args.course == "CSC301"
    assert args.total == 100
    assert args.out == Path("report.json")
    assert args.mode == "MARK"
    assert args.language == "English"
    assert args.review_threshold == 3
    assert args.dpi == 300
    assert args.course_title is None
    assert args.paper is None
    assert args.func is cli.cmd_mark


def test_parser_missing_required_argument_exits_nonzero():
    parser = cli.build_parser()

    with pytest.raises(SystemExit) as exc_info:
        parser.parse_args(["mark", "--scheme", "scheme.txt", "--course", "CSC301"])

    assert exc_info.value.code != 0


def test_help_exits_zero(capsys):
    parser = cli.build_parser()

    with pytest.raises(SystemExit) as exc_info:
        parser.parse_args(["--help"])

    assert exc_info.value.code == 0
    assert "mark" in capsys.readouterr().out


def test_mark_help_exits_zero(capsys):
    parser = cli.build_parser()

    with pytest.raises(SystemExit) as exc_info:
        parser.parse_args(["mark", "--help"])

    assert exc_info.value.code == 0
    assert "--pdf" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# full flow
# ---------------------------------------------------------------------------


def test_mark_command_writes_json_txt_and_excel_on_success(sample_pdf, scheme_file, tmp_path, capsys):
    out_path = tmp_path / "out" / "report.json"
    client = FakeAnthropic([_text_response(_full_response_text(_valid_report_dict()))])

    exit_code = cli.main(
        [
            "mark",
            "--pdf",
            str(sample_pdf),
            "--scheme",
            str(scheme_file),
            "--course",
            "CSC301",
            "--total",
            "10",
            "--out",
            str(out_path),
        ],
        client=client,
    )

    assert exit_code == 0

    assert out_path.exists()
    report_data = json.loads(out_path.read_text(encoding="utf-8"))
    assert report_data["course_code"] == "CSC301"
    assert report_data["student_name"] == "Jane Doe"

    txt_path = out_path.with_suffix(".txt")
    assert txt_path.exists()
    assert "STUDENT EXAMINATION REPORT" in txt_path.read_text(encoding="utf-8")

    xlsx_path = out_path.with_suffix(".xlsx")
    assert xlsx_path.exists()
    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["Report", "Flags"]

    assert "Marked successfully" in capsys.readouterr().out


def test_mark_command_writes_error_payload_and_returns_1_on_failure(
    sample_pdf, scheme_file, tmp_path, capsys
):
    broken_text = "no json here at all"
    client = FakeAnthropic([_text_response(broken_text), _text_response(broken_text)])
    out_path = tmp_path / "report.json"

    exit_code = cli.main(
        [
            "mark",
            "--pdf",
            str(sample_pdf),
            "--scheme",
            str(scheme_file),
            "--course",
            "CSC301",
            "--total",
            "10",
            "--out",
            str(out_path),
        ],
        client=client,
    )

    assert exit_code == 1
    assert out_path.exists()
    payload = json.loads(out_path.read_text(encoding="utf-8"))
    assert payload["success"] is False
    assert payload["raw_output"] == broken_text
    assert not out_path.with_suffix(".txt").exists()
    assert not out_path.with_suffix(".xlsx").exists()
    assert "FAILED" in capsys.readouterr().err
